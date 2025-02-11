import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from descope.descope_client import DescopeClient
from descope.exceptions import AuthException

# Authentication Setup
DESCOPE_PROJECT_ID = str(st.secrets.get("DESCOPE_PROJECT_ID"))
descope_client = DescopeClient(project_id=DESCOPE_PROJECT_ID)

if "token" not in st.session_state:
    if "code" in st.query_params:
        code = st.query_params["code"]
        st.query_params.clear()
        try:
            with st.spinner("Loading..."):
                jwt_response = descope_client.sso.exchange_token(code)
            st.session_state["token"] = jwt_response["sessionToken"].get("jwt")
            st.session_state["refresh_token"] = jwt_response["refreshSessionToken"].get("jwt")
            st.session_state["user"] = jwt_response["user"]
            st.rerun()
        except AuthException:
            st.error("Login failed!")
    st.warning("You're not logged in, please log in.")
    with st.container(border=True):
        if st.button("Sign In with Google", use_container_width=True):
            oauth_response = descope_client.oauth.start(
                provider="google", return_url="https://excelmerged.streamlit.app/"
            )
            url = oauth_response["url"]
            st.markdown(f'<meta http-equiv="refresh" content="0; url={url}">', unsafe_allow_html=True)
else:
    try:
        with st.spinner("Loading..."):
            jwt_response = descope_client.validate_and_refresh_session(
                st.session_state.token, st.session_state.refresh_token
            )
            st.session_state["token"] = jwt_response["sessionToken"].get("jwt")
        st.title("Excel Merger with Authentication")
        st.subheader("Welcome, you're logged in!")
        if "user" in st.session_state:
            user = st.session_state.user
            st.write(f"Name: {user['name']}")
        if st.button("Logout"):
            del st.session_state.token
            st.rerun()
        
        # Excel Merging Application
        st.write("Upload your Excel files to merge them into a single file.")
        key_variable = st.text_input("Enter the key variable for merging:", "")
        sheet_name = st.text_input("Enter the sheet name (optional):", "")
        uploaded_files = st.file_uploader("Upload Excel files", type=["xls", "xlsx"], accept_multiple_files=True)

        if uploaded_files:
            st.write("Uploaded Files:")
            for file in uploaded_files:
                st.write(file.name)

        def sheet_exists(file, sheet_name):
            try:
                wb = load_workbook(file, read_only=True)
                return sheet_name in wb.sheetnames
            except Exception:
                return False

        if sheet_name:
            for file in uploaded_files:
                if not sheet_exists(file, sheet_name):
                    st.warning(f"Sheet '{sheet_name}' does not exist in '{file.name}'")
                    break

        if st.button("Merge"):
            if uploaded_files:
                all_dataframes = []
                variable_names = None
                for file in uploaded_files:
                    if key_variable == "":
                        df = pd.read_excel(file, sheet_name=sheet_name if sheet_name else 0)
                    else:
                        df = pd.read_excel(file)
                    if key_variable == "" and variable_names is None:
                        variable_names = set(df.columns)
                    elif key_variable == "" and variable_names != set(df.columns):
                        st.warning("Files have different variables; merging may result in unexpected data.")
                        break
                    all_dataframes.append(df)
                else:
                    merged_dataframe = pd.concat(all_dataframes) if key_variable == "" else all_dataframes[0]
                    for i in range(1, len(all_dataframes)):
                        merged_dataframe = pd.merge(merged_dataframe, all_dataframes[i], on=key_variable, how="outer")
                    st.write("Merged Dataframe:")
                    st.write(merged_dataframe)
                    merged_dataframe.to_excel("merged_file.xlsx", index=False)
                    st.download_button("Download Merged File", open("merged_file.xlsx", 'rb').read(), "merged_file.xlsx", "application/vnd.ms-excel")
            else:
                st.warning("Please upload at least one file.")
    except AuthException:
        del st.session_state.token
        st.rerun()
